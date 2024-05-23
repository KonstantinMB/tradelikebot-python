import sys
import datetime
import requests
from urllib.parse import urlencode
import time
import hmac
import hashlib
import openpyxl
import pandas as pd
from enum import Enum
import numpy as np

# Feeding auth
from dotenv import load_dotenv
import os
load_dotenv()

symbols = []
CONFIG_PATH = os.getenv('CONFIG_PATH')

#Strategy Parameters
ema_short_period = 45
ema_long_period = 100
regime_filter = 480

# BUY Order Parameters
buy_symbol_pair_order_counter = {}
buy_symbol_pair_order_id = {}
buy_symbol_pair_order_status = {}

# Take Profit Parameters
buy_symbol_pair_target_counter = {}
buy_symbol_pair_target_order_id = {}
buy_symbol_pair_target_order_count = {}
buy_symbol_pair_target_order_status = {}

class OrderStatus(Enum):
    OPEN_ORDER = 'OPEN_ORDER',
    POSITION = 'POSITION'

class TakeProfitStatus(Enum):
    NOT_PLACED = "NOT_PLACED",
    PLACED = 'PLACED'

class Binance():

    apikey = ""
    secretkey = ""
    test = False
    baseurl = ""
    tick_sizes = {}
    qty_steps = {}
    recv_window = 59999
    symbol_base = {}  # base/quote
    symbol_quote = {}

    def __init__(self, apikey="", secretkey="", test=False):

        global CONFIG_PATH
        self.apikey=apikey
        self.secretkey=secretkey
        if test:
            self.baseurl = "https://testnet.binance.vision"
        else:
            self.baseurl = "https://api.binance.com"

        raw = self.getExchangeInfo()

        for r in raw["symbols"]:

            symbol = r["symbol"]
            self.symbol_quote[symbol] = r["quoteAsset"]
            self.symbol_base[symbol] = r["baseAsset"]

            tick_size = 0
            qty_step = 0

            for f in r["filters"]:
                if f["filterType"] == "PRICE_FILTER":
                    tick_size = float(f["tickSize"])
                elif f["filterType"] == "LOT_SIZE":
                    qty_step = float(f["stepSize"])

            self.tick_sizes[symbol] = tick_size
            self.qty_steps[symbol] = qty_step

    def priceRound(self, symbol, price):
        return round(self.tick_sizes[symbol] * int(price / self.tick_sizes[symbol] + 0.5), 9)

    def qtyRound(self, symbol, qty):
        return round(self.qty_steps[symbol] * int(qty / self.qty_steps[symbol] + 0.5), 9)

    def qtyRoundDown(self, symbol, qty):
        return round(self.qty_steps[symbol] * int(qty / self.qty_steps[symbol]), 9)

    def dispatch_request(self, http_method):
        session = requests.Session()
        session.headers.update({
            'Content-Type': 'application/json;charset=utf-8',
            'X-MBX-APIKEY': self.apikey
        })
        return {
            'GET': session.get,
            'DELETE': session.delete,
            'PUT': session.put,
            'POST': session.post,
        }.get(http_method, 'GET')

    def get_timestamp(self):
        return int(time.time() * 1000)

    def hashing(self, query_string):
        return hmac.new(self.secretkey.encode('utf-8'), query_string.encode('utf-8'), hashlib.sha256).hexdigest()

    def send_signed_request(self, http_method, url_path, payload={}):

        payload["recvWindow"] = self.recv_window

        query_string = urlencode(payload, True)
        if query_string:
            query_string = "{}&timestamp={}".format(query_string, self.get_timestamp())
        else:
            query_string = 'timestamp={}'.format(self.get_timestamp())

        url = self.baseurl + url_path + '?' + query_string + '&signature=' + self.hashing(query_string)

        params = {'url': url, 'params': {}}
        response = self.dispatch_request(http_method)(**params)

        return response.json()

    def send_public_request(self, url_path, payload={}):

        query_string = urlencode(payload, True)
        url = self.baseurl + url_path

        if query_string:
            url = url + '?' + query_string

        response = self.dispatch_request('GET')(url=url)
        return response.json()

    def Account(self):
        return self.send_signed_request("GET", "/api/v3/account")

    def getChart(self, symbol, interval, start_t=0, end_t=0):

        if not interval in ["1m", "3m", "5m", "15m", "30m", "1h", "2h", "4h", "6h", "8h", "12h", "1d", "3d", "1w",
                            "1M"]:
            print("getChart: invalid interval")
            return

        payload = {"symbol": symbol, "interval": interval}
        if start_t != 0:
            payload["startTime"] = int(start_t.timestamp() * 1000)
        if end_t != 0:
            payload["endTime"] = int(end_t.timestamp() * 1000)

        #payload = {"symbol":symbol,"interval":interval,"startTime":int(sT.timestamp()*1000),"endTime":int(eT.timestamp()*1000)}
        raw = self.send_public_request("/api/v3/klines", payload=payload)
        chart = {"t": [], "o": [], "h": [], "l": [], "c": [], "v": []}

        for r in raw:
            chart["t"].append(datetime.datetime.fromtimestamp(r[0] / 1000, datetime.timezone.utc))
            chart["o"].append(float(r[1]))
            chart["h"].append(float(r[2]))
            chart["l"].append(float(r[3]))
            chart["c"].append(float(r[4]))
            chart["v"].append(float(r[5]))

        # Creating DataFrame from raw data
        columns = ['Time', 'Open', 'High', 'Low', 'Close', 'Volume', 'Close Time',
                   'Quote Asset Volume', 'Number of Trades', 'Taker Buy Base Asset Volume',
                   'Taker Buy Quote Asset Volume', 'Ignore']
        df = pd.DataFrame(raw, columns=columns)
        df['Time'] = pd.to_datetime(df['Time'], unit='ms')
        df.set_index('Time', inplace=True)
        df = df[['Open', 'High', 'Low', 'Close', 'Volume']].astype(float)

        return chart, df

    def getExchangeInfo(self):
        raw = self.send_public_request("/api/v3/exchangeInfo")
        return raw

    def checkOrder(self, symbol, orderId):

        payload = {"symbol": symbol, "orderId": orderId}
        a = self.send_signed_request("GET", "/api/v3/order", payload=payload)
        # "status": "FILLED" / "NEW" / "CANCELED" / "EXPIRED"
        return a

    def Buy(self, symbol, quantity, price):
        payload = {"symbol": symbol, "side": "BUY", "quantity": self.qtyRound(symbol, quantity)}
        if price == 0:
            payload["type"] = "MARKET"
        else:
            payload["type"] = "LIMIT"
            payload["timeInForce"] = "GTC"
            payload["price"] = self.priceRound(symbol, price)
        #print(payload)
        a = self.send_signed_request("POST", "/api/v3/order", payload=payload)

        #print(a)
        if not "orderId" in a:
            std_log(str(a))
            std_log(payload)
            std_log("Order error")
            sys.exit()
        std_log("[%s] Buy (quantity:%f, price:%f, orderID:%d)" % (symbol, quantity, price, a["orderId"]))

        update_buy_order_id_for_symbol_pair(a, symbol)
        buy_symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER
        return a

    def ReplaceOrder(self, order_id, symbol_pair, quantity, price):

        payload = {
            "symbol": symbol_pair,
            "side": 'BUY',
            "type": "LIMIT",
            "timeInForce": "GTC",
            "cancelReplaceMode": "STOP_ON_FAILURE",
            "cancelOrderId": order_id,
            "quantity": self.qtyRound(symbol, quantity),
            "price": self.priceRound(symbol, price)
        }

        try:
            replaced_order_response = (
                self.send_signed_request("POST", "/api/v3/order/cancelReplace", payload))

            std_log("[%s] BUY Order Replaced. New Order Parameters: (quantity:%f, price:%f, orderID:%d)"
                    % (symbol, quantity, price, replaced_order_response["newOrderResponse"]["orderId"]))

            update_buy_order_id_for_symbol_pair(replaced_order_response["newOrderResponse"], symbol)
            buy_symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER

            return replaced_order_response["newOrderResponse"]
        except Exception as e:
            std_log(f"[{symbol_pair}] Error canceling order {order_id}. Error Info: {e}")
            return None

    def ReplaceTakeProfitOrder(self, order_id, symbol_pair, buy_amount, new_price):

        # Prepare the payload for the limit sell order
        payload = {
            "symbol": symbol_pair,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "cancelReplaceMode": "STOP_ON_FAILURE",
            "cancelOrderId": order_id,
            "quantity": self.qtyRound(symbol, buy_amount),
            "price": self.priceRound(symbol, new_price)
        }

        try:
            replaced_order_response = (
                self.send_signed_request("POST", "/api/v3/order/cancelReplace", payload))
            return replaced_order_response["newOrderResponse"]
        except Exception as e:
            std_log(f"[{symbol_pair}] Error canceling order {order_id}. Error Info: {e}")
            return None

    def GetContractPrice(self, symbol, orderId):
        payload = {"symbol": symbol}
        a = self.send_signed_request("GET", "/api/v3/myTrades", payload=payload)
        price = -1
        for order in a:
            if order["orderId"] == int(orderId):
                price = float(order["price"])
                break
        return price

    def boundaryRemaining(self, tf):

        # "1m", "3m", "5m", "15m", "30m", "1h", "2h", "4h", "6h", "8h", "12h", "1d", "3d", "1w", "1M"
        t = datetime.datetime.now(datetime.timezone.utc)
        if tf == "1m":
            next_t = (t + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
        elif tf == "3m":
            next_t = (t + datetime.timedelta(minutes=3 - t.minute % 3)).replace(second=0, microsecond=0)
        elif tf == "5m":
            next_t = (t + datetime.timedelta(minutes=5 - t.minute % 5)).replace(second=0, microsecond=0)
        elif tf == "15m":
            next_t = (t + datetime.timedelta(minutes=15 - t.minute % 15)).replace(second=0, microsecond=0)
        elif tf == "30m":
            next_t = (t + datetime.timedelta(minutes=30 - t.minute % 30)).replace(second=0, microsecond=0)
        elif tf == "1h":
            next_t = (t + datetime.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
        elif tf == "2h":
            next_t = (t + datetime.timedelta(hours=2 - t.hour % 2)).replace(minute=0, second=0, microsecond=0)
        elif tf == "4h":
            next_t = (t + datetime.timedelta(hours=4 - t.hour % 4)).replace(minute=0, second=0, microsecond=0)
        elif tf == "6h":
            next_t = (t + datetime.timedelta(hours=6 - t.hour % 6)).replace(minute=0, second=0, microsecond=0)
        elif tf == "8h":
            next_t = (t + datetime.timedelta(hours=8 - t.hour % 8)).replace(minute=0, second=0, microsecond=0)
        elif tf == "12h":
            next_t = (t + datetime.timedelta(hours=12 - t.hour % 12)).replace(minute=0, second=0, microsecond=0)
        elif tf == "1d":
            next_t = (t + datetime.timedelta(hours=24)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif tf == "3d":
            day_pivot = datetime.datetime(2017, 8, 17, 0, 0)
            next_t = (t + datetime.timedelta(days=3 - (t - day_pivot).days % 3)).replace(hour=0, minute=0, second=0,
                                                                                         microsecond=0)
        elif tf == "1w":  # Monday 0am
            next_t = (t + datetime.timedelta(days=7 - t.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        elif tf == "1M":  # 0am
            if t.month == 12:
                next_t = datetime.datetime(t.year + 1, 1, 1, 0, 0, 0)
            else:
                next_t = datetime.datetime(t.year, t.month + 1, 1, 0, 0, 0)
        remaining = next_t - t
        return remaining

    def replace_position_with_new_order(self, symbol_pair, order_id, buy_amount, new_price):
        try:
            self.ReplaceOrder(order_id, symbol_pair, buy_amount, new_price)
        except Exception as e:
            std_log(f"[{symbol_pair}] Error modifying order. Error Info: {e}")
            return None

    def set_take_profit(self, symbol_pair, quantity, price):

        """
               Place a limit order to sell a cryptocurrency at a specified price.

               Args:
               symbol (str): The trading pair symbol, e.g., 'BTCUSDT'.
               buy_amount (float): The amount of the cryptocurrency to sell.
               price (float): The price at which the order should execute.
               """

        # Prepare the payload for the limit sell order
        payload = {
            "symbol": symbol_pair,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",  # Good till cancelled
            "quantity": self.qtyRound(symbol, quantity),
            "price": self.priceRound(symbol, price)
        }

        try:
            response = self.send_signed_request("POST", "/api/v3/order", payload)
            if response and 'orderId' in response:
                update_take_profit_order_id_for_symbol_pair(response, symbol_pair)
                buy_symbol_pair_target_order_status[symbol] = TakeProfitStatus.PLACED
                std_log(f"[{symbol_pair}] Target Order Placed. Will Sell {quantity} of {symbol_pair} at: {price}!"
                        f" Order ID: {response['orderId']}")
                return response
            else:
                std_log(f"[{symbol_pair}] Failed to place target order. Response: {response}")
                return None
        except Exception as e:
            std_log(f"[{symbol_pair}] Error placing target order. Error Info: {e}")
            return None

    def update_take_profit(self, symbol_pair, order_id, quantity, new_take_profit_price):

        try:

            response = self.ReplaceTakeProfitOrder(order_id, symbol_pair, quantity, new_take_profit_price)

            if response and 'orderId' in response:

                update_take_profit_order_id_for_symbol_pair(response, symbol_pair)
                buy_symbol_pair_target_order_status[symbol_pair] = TakeProfitStatus.PLACED

                std_log("[%s] Target Order Replaced. New Order Parameters: (quantity:%f, price:%f, orderID:%d)"
                        % (symbol_pair, quantity, new_take_profit_price, response["orderId"]))

                return response

            else:
                std_log(f"[{symbol_pair}] Failed to place new take profit order. Response: {response}")
                return None

        except Exception as e:
            std_log(f"[{symbol_pair}] Error placing take_profit for order. Error Info: {e}")
            return None

    def check_order_status(self, symbol_pair, order_id):

        try:

            response = self.checkOrder(symbol=symbol_pair, orderId=order_id)

            if response:
                order_status = response
                std_log(f"[{symbol_pair}] Order ID [{order_id}] with status: [{order_status['status']}]")
                return order_status['status']
            else:
                std_log(f"[{symbol_pair}] No response received for order status request.")
                return None

        except Exception as e:
            std_log(f"[{symbol_pair}] Error checking order status. Error Info: {str(e)}")
            return None

    def check_target_order_status(self, symbol_pair, order_id):

        try:

            response = self.checkOrder(symbol=symbol_pair, orderId=order_id)

            if response:
                order_status = response
                std_log(f"[{symbol_pair}] Target Order ID [{order_id}] with status: [{order_status['status']}]")
                return order_status['status']
            else:
                std_log(f"[{symbol_pair}] No response received for target order status request.")
                return None

        except Exception as e:
            std_log(f"[{symbol_pair}] Error checking target order status. Error Info: {str(e)}")
            return None

def configure_parameters(excel_path):

    buy_timedelta = {}
    buy_timeframe = {}
    buy_order_type = {}
    order_size = {}
    h_period = {}
    demo = True
    buy_limit = 0
    tdelta_conv = {"1m": datetime.timedelta(minutes=1), "3m": datetime.timedelta(minutes=3),
                   "5m": datetime.timedelta(minutes=5), "15m": datetime.timedelta(minutes=15),
                   "30m": datetime.timedelta(minutes=30), "1h": datetime.timedelta(hours=1),
                   "2h": datetime.timedelta(hours=2), "4h": datetime.timedelta(hours=4),
                   "6h": datetime.timedelta(hours=6), "8h": datetime.timedelta(hours=8),
                   "12h": datetime.timedelta(hours=12), "1d": datetime.timedelta(days=1),
                   "3d": datetime.timedelta(days=3), "1w": datetime.timedelta(days=7)}

    # 0.1. Read excel
    config = openpyxl.load_workbook(excel_path)
    sheet = config.worksheets[0]

    for row_n in range(2, 100):
        if sheet.cell(row=row_n, column=1).value == "END OF CONFIGURATION":
            break
        if sheet.cell(row=row_n, column=3).value == "BUY":
            symbol = sheet.cell(row=row_n, column=1).value
            if not symbol in symbols:
                symbols.append(symbol)
            buy_timeframe[symbol] = sheet.cell(row=row_n, column=4).value
            buy_order_type[symbol] = sheet.cell(row=row_n, column=5).value
            order_size[symbol] = float(sheet.cell(row=row_n, column=6).value)
            h_period[symbol] = int(sheet.cell(row=row_n, column=7).value)
        elif sheet.cell(row=row_n, column=1).value == "Open positions limit":
            buy_limit = int(sheet.cell(row=row_n, column=2).value)
        elif sheet.cell(row=row_n, column=1).value == "Demo trading":
            demo = sheet.cell(row=row_n, column=2).value

    std_log("[Booting] Buy timeframes %s" % str(buy_timeframe))
    std_log("[Booting] Buy order type %s" % str(buy_order_type))
    std_log("[Booting] Order size %s" % str(order_size))
    std_log("[Booting] Highest period %s" % str(h_period))
    std_log("[Booting] Open positions limit %s" % str(buy_limit))
    std_log("[Booting] Demo account: %s" % str(demo))

    for symbol_pair in symbols:
        if symbol_pair in buy_timeframe:
            buy_timedelta[symbol_pair] = tdelta_conv[buy_timeframe[symbol_pair]]  # to get candle closing period
        else:
            buy_timedelta[symbol_pair] = None  # or some default value, or skip setting it altogether

    return buy_timedelta, buy_timeframe, buy_order_type, order_size, h_period, demo, buy_limit

def configure_binance_client():

    test_api_key = os.getenv('TEST_API_KEY')
    test_api_secret = os.getenv('TEST_API_SECRET')
    api_key = os.getenv('API_KEY')
    api_secret = os.getenv('API_SECRET')

    # Legacy Client Initialization
    binance = Binance(test=demo, apikey='', secretkey='')

    if demo:
        binance = Binance(test=demo, apikey=test_api_key, secretkey=test_api_secret)
    else:
        binance = Binance(test=demo, apikey=api_key, secretkey=api_secret)

    return binance

def set_old_remain_buy():

    old_remain_buy = {}
    for symbol in symbols:
        old_remain_buy[symbol] = datetime.timedelta(days=365)  # to get candle closing period

    return old_remain_buy

def custom_Nate_conditions(order_size):

    # Multi-timeframe analysis for daily data - Adding conditions to filter for daily uptrends

    chart_d, chart_df_d = binance.getChart(symbol, "1d",  start_t=datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=365))

    close_p = chart_d["c"][-1]
    quantity_p = order_size[symbol] / close_p

    # Calculate the short-term (12-period) exponential moving average (EMA)
    short_ema = chart_df['Close'].ewm(span=12, adjust=False).mean()

    # Calculate the long-term (26-period) exponential moving average (EMA)
    long_ema = chart_df['Close'].ewm(span=26, adjust=False).mean()

    # Calculate the MACD line
    macd_line = short_ema - long_ema

    # Calculate the signal line (9-period EMA of the MACD line)
    signal_line = macd_line.ewm(span=9, adjust=False).mean()

    # Calculate the MACD histogram
    macd_histogram = macd_line - signal_line
    # Add MACD indicators to the DataFrame
    chart_df_d['HHV_56'] = chart_df_d['High'].rolling(window=56).max()
    chart_df_d['MACD Line'] = macd_line
    chart_df_d['Signal Line'] = signal_line
    chart_df_d['MACD Histogram'] = macd_histogram

    # Add a column indicating whether MACD histogram is above 0 or not
    chart_df_d['MACD Above 0'] = chart_df_d['MACD Histogram'] > 0
    #last_day_data = chart_df_d.iloc[-1]

    #AND (BBandTop_/BBandBot_) -1 > MinBands   AND  Hist_d_stock > 0 AND C > from_peak * period_high_week  AND   C > MA(C,480) AND EMA(C,SEMAsFilter) > EMA(C,LEMAsFilter)

    #Here I will add all the necessary conditions from the strategy backtest: Bollinger Bands, EMA, etc.
    conditions_met = chart_df["upper_band"].iloc[-1]/chart_df["lower_band"].iloc[-1] > 1 and \
                     chart_df["ema_regime_filter"].iloc[-1] > -1 and \
                     chart_df["ema_trend_filter"].iloc[-1] > -1 and \
                     chart_df_d['MACD Histogram'].iloc[-1] > -2 and \
                     chart_df['Close'].iloc[-1] > chart_df_d["HHV_56"].iloc[-1]* 0.30

    return 1 # to overwrite and not use conditions in main logic

def get_chart_data(order_size, symbol, interval, start_t):

    candle_chart, candle_chart_df = binance.getChart(symbol, interval, start_t=start_t)

    if candle_chart["t"][-1].astimezone(datetime.timezone.utc) > datetime.datetime.now(datetime.timezone.utc) - \
            buy_timedelta[symbol] / 2:
        candle_chart["t"] = candle_chart["t"][:-1]  # Remove excessive candle
        candle_chart["o"] = candle_chart["o"][:-1]
        candle_chart["h"] = candle_chart["h"][:-1]
        candle_chart["l"] = candle_chart["l"][:-1]
        candle_chart["c"] = candle_chart["c"][:-1]
        candle_chart["v"] = candle_chart["v"][:-1]

    # Ensure DataFrame is not empty and is sorted by index (Open Time)
    if candle_chart_df.empty or candle_chart_df.index[-1].tz_localize('UTC') > datetime.datetime.now(
            datetime.timezone.utc) - buy_timedelta[symbol] / 2:
        # Remove the last row if the candle is not fully formed
        candle_chart_df = candle_chart_df.iloc[:-1]

    candle_chart_df["rolling_mean"] = candle_chart_df['Close'].rolling(window=20).mean()
    candle_chart_df["rolling_std"] = candle_chart_df['Close'].rolling(window=20).std()
    candle_chart_df["upper_band"] = candle_chart_df["rolling_mean"] + (2 * candle_chart_df["rolling_std"])
    candle_chart_df["lower_band"] = candle_chart_df["rolling_mean"] - (2 * candle_chart_df["rolling_std"])
    candle_chart_df["ema_short_period"] = candle_chart_df['Close'].ewm(span=ema_short_period).mean()
    candle_chart_df["ema_long_period"] = candle_chart_df['Close'].ewm(span=ema_long_period).mean()
    candle_chart_df["ema_regime"] =  candle_chart_df['Close'].ewm(span=regime_filter).mean()

    candle_chart_df["ema_trend_filter"] = candle_chart_df["ema_short_period"] > candle_chart_df["ema_long_period"]
    candle_chart_df["ema_regime_filter"] = candle_chart_df["Close"] > candle_chart_df["ema_regime"]
    candle_chart_df["bbands_width_filter"] = np.where(candle_chart_df["upper_band"]/candle_chart_df["lower_band"] > 1.03, True, False)

    close_p = candle_chart["c"][-1]
    quantity = order_size[symbol] / close_p

    # Calculate the short-term (12-period) exponential moving average (EMA)
    short_ema = candle_chart_df['Close'].ewm(span=12, adjust=False).mean()

    # Calculate the long-term (26-period) exponential moving average (EMA)
    long_ema = candle_chart_df['Close'].ewm(span=26, adjust=False).mean()

    # Calculate the MACD line
    macd_line = short_ema - long_ema

    # Calculate the signal line (9-period EMA of the MACD line)
    signal_line = macd_line.ewm(span=9, adjust=False).mean()

    # Calculate the MACD histogram
    macd_histogram = macd_line - signal_line

    # Add MACD indicators to the DataFrame
    candle_chart_df['MACD Line'] = macd_line
    candle_chart_df['Signal Line'] = signal_line
    candle_chart_df['MACD Histogram'] = macd_histogram

    # Add a column indicating whether MACD histogram is above 0 or not
    candle_chart_df['MACD Above 0'] = candle_chart_df['MACD Histogram'] > 0

    # Add a column indicating crossover points of MACD and Signal lines
    candle_chart_df['MACD_Crossover'] = np.where((candle_chart_df['MACD Line'] > candle_chart_df['Signal Line']) & (
            candle_chart_df['MACD Line'].shift(1) < candle_chart_df['Signal Line'].shift(1)), 1,
                                          np.where((candle_chart_df['MACD Line'] < candle_chart_df['Signal Line']) & (
                                                  candle_chart_df['MACD Line'].shift(1) > candle_chart_df[
                                              'Signal Line'].shift(1)), -1, 0))

    return candle_chart, candle_chart_df, quantity

def reset_dict_for_symbol(symbol):

    buy_symbol_pair_order_counter[symbol] = 0
    buy_symbol_pair_order_id[symbol] = 0
    buy_symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER

    buy_symbol_pair_target_counter[symbol] = 0
    buy_symbol_pair_target_order_id[symbol] = 0
    buy_symbol_pair_target_order_count[symbol] = 0
    buy_symbol_pair_target_order_status[symbol] = TakeProfitStatus.NOT_PLACED

def set_dicts(symbols):

    global buy_symbol_pair_order_counter, buy_symbol_pair_order_id, buy_symbol_pair_target_order_count, \
        buy_symbol_pair_target_counter, buy_symbol_pair_target_order_id, buy_symbol_pair_order_status, \
        buy_symbol_pair_order_status, buy_symbol_pair_target_order_status

    buy_symbol_pair_order_counter = {symbol: 0 for symbol in symbols}
    buy_symbol_pair_order_id = {symbol: 0 for symbol in symbols}
    buy_symbol_pair_order_status = {symbol: OrderStatus.OPEN_ORDER for symbol in symbols}

    buy_symbol_pair_target_counter = {symbol: 0 for symbol in symbols}
    buy_symbol_pair_target_order_id = {symbol: 0 for symbol in symbols}
    buy_symbol_pair_target_order_count = {symbol: 0 for symbol in symbols}
    buy_symbol_pair_target_order_status = {symbol: TakeProfitStatus.NOT_PLACED for symbol in symbols}

def cur_time():
    s = "[" + time.strftime("%d%b%Y", time.localtime()) + "]"
    s = s + "[" + time.strftime("%H:%M:%S", time.localtime()) + "]"
    return s.upper()

def std_log(s):
    global CONFIG_PATH
    s = str(s)
    print(cur_time() + s)
    fout = open(CONFIG_PATH + "log%s.txt" % (time.strftime("%y%m%d", time.localtime())), "a")
    fout.writelines(cur_time() + s + "\n")
    fout.close()

def update_take_profit_order_id_for_symbol_pair(order_info, symbol_pair):

    global buy_symbol_pair_target_order_id
    if order_info is not None:
        buy_symbol_pair_target_order_id[symbol_pair] = order_info['orderId']

    return buy_symbol_pair_target_order_id[symbol_pair]

def update_buy_order_id_for_symbol_pair(order_info, symbol_pair):

    global buy_symbol_pair_order_id
    if order_info is not None:
        buy_symbol_pair_order_id[symbol_pair] = order_info['orderId']

    return buy_symbol_pair_order_id[symbol_pair]

def get_latest_bbands(candles_chart_df: pd.DataFrame):

    closes = candles_chart_df['Close'].values
    dates = pd.to_datetime(candles_chart_df.index)
    close_prices = pd.Series(closes, index=dates)

    # Calculate Bollinger Bands
    rolling_mean = close_prices.rolling(window=20).mean()
    rolling_std = close_prices.rolling(window=20).std()
    upper_band = rolling_mean + (2 * rolling_std)
    lower_band = rolling_mean - (2 * rolling_std)

    latest_close_price = closes[-1]
    latest_upper_band_price = upper_band.iloc[-1]
    latest_lower_band_price = lower_band.iloc[-1]

    return latest_close_price, latest_lower_band_price, latest_upper_band_price

def check_bband_buy_signal(symbol_pair, latest_close, latest_lower_bband_price):

    # Check if the latest close price is below or equal to the Lower Bollinger Band
    if latest_close <= latest_lower_bband_price:
        std_log(f"[{symbol_pair}] Latest Close Price {latest_close} is below or equal to "
                f"the Lower Latest Bollinger Band {latest_lower_bband_price}")
        return False
    else:
        std_log(f"[{symbol_pair}] Bollinger Bands filter condition met "
                f"(Latest Lower Bollinger Band: {latest_lower_bband_price} / Latest Close: {latest_close})")
        return True


if __name__ == "__main__":

    (buy_timedelta,
     buy_timeframe,
     buy_order_type,
     order_size,
     h_period,
     demo,
     buy_limit) = configure_parameters(CONFIG_PATH + "/Bot_config.xlsx")

    binance = configure_binance_client()

    old_remain_buy = set_old_remain_buy()

    set_dicts(symbols)

    std_log("[Booting] Complete")

    for symbol in symbols:
        (chart, chart_df, quantity) = get_chart_data(order_size,
                                                     symbol,
                                                     buy_timeframe[symbol],
                                                     start_t=datetime.datetime.now(datetime.timezone.utc) -
                                                             buy_timedelta[symbol]
                                                             * h_period[symbol] * 2)
        (latest_close_price,
         latest_lower_bband_price,
         latest_upper_bband_price) = get_latest_bbands(chart_df)

        take_profit_response = binance.set_take_profit(symbol,
                                                           quantity,
                                                           latest_upper_bband_price)

        time.sleep(5)
        take_profit_order_id = buy_symbol_pair_target_order_id[symbol]

        take_profit_status = binance.check_target_order_status(symbol, take_profit_order_id)
        time.sleep(5)

        binance.update_take_profit(symbol, take_profit_order_id, quantity, latest_upper_bband_price)

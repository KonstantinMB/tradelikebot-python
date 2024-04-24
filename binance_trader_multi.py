#======================== NOTES ========================#
# - This project was written in python3
# - Required package: openpyxl (conda may include openpyxl as default)
# - CONFIG_PATH and EXCEL_NAME need to be specificed (line 291 - 293)
# - API key and Secret Key can be specified in code or imported from external file.
#    - In case of importing, key.txt must exist in CONFIG_PATH, containing API key and Secret key.
# - In CONFIG_PATH, balance.txt and logYYMMDD.txt will be generated automatically.
#    - balance.txt : Contains open position information and used in session recovery.
#    - logYYMMDD.txt : Contains code execution history and needed in debugging.
#=======================================================#


import sys
import datetime
import requests
from urllib.parse import urlencode
import time
import hmac
import hashlib
import openpyxl
import pandas as pd
from enum import Enum
from pandas import DataFrame, Series

# Feeding auth
from dotenv import load_dotenv
import os
load_dotenv()

class OrderStatus(Enum):
    OPEN_ORDER = 'OPEN_ORDER',
    POSITION = 'POSITION'

symbol_pair_trade_counter = {}
symbol_pair_order_count = {}
symbol_pair_target_order_count = {}
symbol_pair_order_status = {}
symbol_pair_target_order_status = {}

symbols = []

def cur_time():
    s = "[" + time.strftime("%d%b%Y", time.localtime()) + "]"
    s = s + "[" + time.strftime("%H:%M:%S", time.localtime()) + "]"
    return s.upper()

def today(length=6):
    if length == 8:
        return time.strftime("%Y%m%d", time.localtime())
    elif length == 6:
        return time.strftime("%y%m%d", time.localtime())
    elif length == 4:
        return time.strftime("%m%d", time.localtime())

def EMA(closes, period):
    mult = 2 / (period + 1)
    ema = [sum(closes[0:period]) / period] * period
    for c in closes[period:]:
        ema.append((c - ema[-1]) * mult + ema[-1])
    return ema

def std_log(s):
    global CONFIG_PATH
    s = str(s)
    print(cur_time() + s)
    fout = open(CONFIG_PATH + "log%s.txt" % (today(6)), "a")
    fout.writelines(cur_time() + s + "\n")
    fout.close()

def update_balance(balance):
    global CONFIG_PATH
    fout = open(CONFIG_PATH + "balance.txt", "w")
    fout.writelines("%s" % balance)
    fout.close()

class Binance():
    apikey = ""
    secretkey = ""
    test = False
    baseurl = ""
    tick_sizes = {}
    qty_steps = {}
    recv_window = 1000000000
    symbol_base = {}  # base/quote
    symbol_quote = {}

    def __init__(self, apikey="", secretkey="", test=False):
        global CONFIG_PATH
        if apikey == "":
            if test:
                fin = open(CONFIG_PATH + "/key_test.txt", "r")
            else:
                fin = open(CONFIG_PATH + "/key.txt", "r")
            self.apikey = fin.readline().strip()
            self.secretkey = fin.readline().strip()
            fin.close()
            if self.apikey == "":
                print("Please fill the API Key in %s/key.txt" % CONFIG_PATH)
                sys.sys.exit()
        else:
            self.apikey = apikey
            self.secretkey = secretkey

        if test:
            self.baseurl = "https://testnet.binance.vision"
        else:
            self.baseurl = "https://api.binance.com"

        raw = self.getExchangeInfo()
        for r in raw["symbols"]:
            symbol = r["symbol"]
            self.symbol_quote[symbol] = r["quoteAsset"]
            self.symbol_base[symbol] = r["baseAsset"]
            #print(symbol)
            tick_size = 0
            qty_step = 0
            for f in r["filters"]:
                if f["filterType"] == "PRICE_FILTER":
                    tick_size = float(f["tickSize"])
                elif f["filterType"] == "LOT_SIZE":
                    qty_step = float(f["stepSize"])
            self.tick_sizes[symbol] = tick_size
            self.qty_steps[symbol] = qty_step

    def getBase(self, symbol):
        return self.symbol_base[symbol]

    def getQuote(self, symbol):
        return self.symbol_quote[symbol]

    def priceRound(self, symbol, price):
        return round(self.tick_sizes[symbol] * int(price / self.tick_sizes[symbol] + 0.5), 9)

    def qtyRound(self, symbol, qty):
        #print(self.qty_steps[symbol])
        return round(self.qty_steps[symbol] * int(qty / self.qty_steps[symbol] + 0.5), 9)

    def priceRoundDown(self, symbol, price):
        return round(self.tick_sizes[symbol] * int(price / self.tick_sizes[symbol]), 9)

    def qtyRoundDown(self, symbol, qty):
        #print(self.qty_steps[symbol])
        return round(self.qty_steps[symbol] * int(qty / self.qty_steps[symbol]), 9)

    def dispatch_request(self, http_method):
        session = requests.Session()
        session.headers.update({
            'Content-Type': 'application/json;charset=utf-8',
            'X-MBX-APIKEY': self.apikey
        })
        return {
            'GET': session.get,
            'DELETE': session.delete,
            'PUT': session.put,
            'POST': session.post,
        }.get(http_method, 'GET')

    def get_timestamp(self):
        return int(time.time() * 1000)

    def hashing(self, query_string):
        return hmac.new(self.secretkey.encode('utf-8'), query_string.encode('utf-8'), hashlib.sha256).hexdigest()

    def send_signed_request(self, http_method, url_path, payload={}):
        payload["recvWindow"] = self.recv_window
        query_string = urlencode(payload, True)
        if query_string:
            query_string = "{}&timestamp={}".format(query_string, self.get_timestamp())
        else:
            query_string = 'timestamp={}'.format(self.get_timestamp())

        url = self.baseurl + url_path + '?' + query_string + '&signature=' + self.hashing(query_string)
        #print("{} {}".format(http_method, url))
        params = {'url': url, 'params': {}}
        response = self.dispatch_request(http_method)(**params)
        return response.json()

    def send_public_request(self, url_path, payload={}):
        query_string = urlencode(payload, True)
        url = self.baseurl + url_path
        if query_string:
            url = url + '?' + query_string
        #print("{}".format(url))
        response = self.dispatch_request('GET')(url=url)
        return response.json()

    def Account(self):
        return self.send_signed_request("GET", "/api/v3/account")

    def getChart(self, symbol, interval, start_t=0, end_t=0):

        if not interval in ["1m", "3m", "5m", "15m", "30m", "1h", "2h", "4h", "6h", "8h", "12h", "1d", "3d", "1w",
                            "1M"]:
            print("getChart: invalid interval")
            return

        payload = {"symbol": symbol, "interval": interval}
        if start_t != 0:
            payload["startTime"] = int(start_t.timestamp() * 1000)
        if end_t != 0:
            payload["endTime"] = int(end_t.timestamp() * 1000)

        #payload = {"symbol":symbol,"interval":interval,"startTime":int(sT.timestamp()*1000),"endTime":int(eT.timestamp()*1000)}
        raw = self.send_public_request("/api/v3/klines", payload=payload)
        chart = {"t": [], "o": [], "h": [], "l": [], "c": [], "v": []}

        for r in raw:
            chart["t"].append(datetime.datetime.utcfromtimestamp(r[0] / 1000))
            #chart["t"].append(datetime.datetime.fromtimestamp(r[0]/1000))
            chart["o"].append(float(r[1]))
            chart["h"].append(float(r[2]))
            chart["l"].append(float(r[3]))
            chart["c"].append(float(r[4]))
            chart["v"].append(float(r[5]))

        # Creating DataFrame from raw data
        columns = ['Time', 'Open', 'High', 'Low', 'Close', 'Volume', 'Close Time',
                   'Quote Asset Volume', 'Number of Trades', 'Taker Buy Base Asset Volume',
                   'Taker Buy Quote Asset Volume', 'Ignore']
        df = pd.DataFrame(raw, columns=columns)
        df['Time'] = pd.to_datetime(df['Time'], unit='ms')
        df.set_index('Time', inplace=True)
        df = df[['Open', 'High', 'Low', 'Close', 'Volume']].astype(float)

        return chart, df

    def getExchangeInfo(self):
        raw = self.send_public_request("/api/v3/exchangeInfo")
        return raw

    def getBook(self, symbol):
        payload = {"symbol": symbol}
        raw = self.send_public_request("/api/v3/depth", payload=payload)
        #print(raw)  # TODO: remove
        return raw

    def checkOrder(self, symbol, orderId):
        payload = {"symbol": symbol, "orderId": orderId}
        a = self.send_signed_request("GET", "/api/v3/order", payload=payload)
        # "status": "FILLED" / "NEW" / "CANCELED" / "EXPIRED"
        return a

    def Buy(self, symbol, quantity, price):
        payload = {"symbol": symbol, "side": "BUY", "quantity": self.qtyRound(symbol, quantity)}
        if price == 0:
            payload["type"] = "MARKET"
        else:
            payload["type"] = "LIMIT"
            payload["timeInForce"] = "GTC"
            payload["price"] = self.priceRound(symbol, price)
        #print(payload)
        a = self.send_signed_request("POST", "/api/v3/order", payload=payload)

        #print(a)
        if not "orderId" in a:
            std_log(str(a))
            std_log(payload)
            std_log("Order error")
            sys.exit()
        std_log("Buy %s (quantity:%f, price:%f, orderID:%d)" % (symbol, quantity, price, a["orderId"]))

        self.update_order_id_for_symbol_pair(a, symbol)
        symbol_pair_trade_counter[symbol] += 1
        symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER
        return a

    def Sell(self, symbol, quantity, price):
        payload = {"symbol": symbol, "side": "SELL", "quantity": self.qtyRound(symbol, quantity)}
        if price == 0:
            payload["type"] = "MARKET"
        else:
            payload["type"] = "LIMIT"
            payload["timeInForce"] = "GTC"
            payload["price"] = self.priceRound(symbol, price)
        a = self.send_signed_request("POST", "/api/v3/order", payload=payload)
        #print(a)
        if not "orderId" in a:
            std_log(str(a))
            std_log(payload)
            std_log("Order error")
            sys.exit()
        std_log("Sell %s (quantity:%f, orderID:%d)" % (symbol, quantity, a["orderId"]))

        symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER
        return a

    def OrderCancel(self, symbol_pair, order_id):

        cancel_payload = {
            "symbol": symbol_pair,
            "orderId": order_id
        }

        try:
            cancel_response = self.send_signed_request("DELETE", "/api/v3/order", cancel_payload)
            if cancel_response and cancel_response.get('status', '') == 'CANCELED':
                std_log(f"[{symbol_pair}] Order {order_id} canceled successfully.")
            else:
                std_log(f"[{symbol_pair}] Failed to cancel order {order_id}. Response: {cancel_response}")
                return None
        except Exception as e:
            std_log(f"[{symbol_pair}] Error canceling order {order_id}. Error Info: {e}")
            return None

    def getBalanceQuantity(self, symbol):

        while True:
            acc = self.Account()
            acc = acc["balances"]
            quantity = -1
            for asset in acc:
                if asset["asset"] == symbol:
                    quantity = float(asset["free"])
                    break
            std_log("%s %f in balance" % (symbol, quantity))
            if quantity != -1: break
            time.sleep(1)
        return quantity

    def OrderWait(self, symbol, orderId):
        #time.sleep(10)
        order_filled = False
        try:
            while True:
                order = self.checkOrder(symbol, orderId)
                # {'code': -2013, 'msg': 'Order does not exist.'}
                if not "status" in order:
                    print(order)
                    print("Warning: Server did not accept order yet.")
                elif order["status"] != "NEW":
                    order_filled = True
                    break
                time.sleep(1)
        except:
            std_log("%s" % str(order))
            std_log("Order %s canceled" % orderId)
            self.OrderCancel(symbol, orderId)
        return order_filled

    def GetContractPrice(self, symbol, orderId):
        payload = {"symbol": symbol}
        a = self.send_signed_request("GET", "/api/v3/myTrades", payload=payload)
        price = -1
        for order in a:
            if order["orderId"] == int(orderId):
                price = float(order["price"])
                break
        return price

    def boundaryRemaining(self, tf):

        # "1m", "3m", "5m", "15m", "30m", "1h", "2h", "4h", "6h", "8h", "12h", "1d", "3d", "1w", "1M"
        t = datetime.datetime.now(datetime.UTC)
        if tf == "1m":
            next_t = (t + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)
        elif tf == "3m":
            next_t = (t + datetime.timedelta(minutes=3 - t.minute % 3)).replace(second=0, microsecond=0)
        elif tf == "5m":
            next_t = (t + datetime.timedelta(minutes=5 - t.minute % 5)).replace(second=0, microsecond=0)
        elif tf == "15m":
            next_t = (t + datetime.timedelta(minutes=15 - t.minute % 15)).replace(second=0, microsecond=0)
        elif tf == "30m":
            next_t = (t + datetime.timedelta(minutes=30 - t.minute % 30)).replace(second=0, microsecond=0)
        elif tf == "1h":
            next_t = (t + datetime.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
        elif tf == "2h":
            next_t = (t + datetime.timedelta(hours=2 - t.hour % 2)).replace(minute=0, second=0, microsecond=0)
        elif tf == "4h":
            next_t = (t + datetime.timedelta(hours=4 - t.hour % 4)).replace(minute=0, second=0, microsecond=0)
        elif tf == "6h":
            next_t = (t + datetime.timedelta(hours=6 - t.hour % 6)).replace(minute=0, second=0, microsecond=0)
        elif tf == "8h":
            next_t = (t + datetime.timedelta(hours=8 - t.hour % 8)).replace(minute=0, second=0, microsecond=0)
        elif tf == "12h":
            next_t = (t + datetime.timedelta(hours=12 - t.hour % 12)).replace(minute=0, second=0, microsecond=0)
        elif tf == "1d":
            next_t = (t + datetime.timedelta(hours=24)).replace(hour=0, minute=0, second=0, microsecond=0)
        elif tf == "3d":
            day_pivot = datetime.datetime(2017, 8, 17, 0, 0)
            next_t = (t + datetime.timedelta(days=3 - (t - day_pivot).days % 3)).replace(hour=0, minute=0, second=0,
                                                                                         microsecond=0)
        elif tf == "1w":  # Monday 0am
            next_t = (t + datetime.timedelta(days=7 - t.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        elif tf == "1M":  # 0am
            if t.month == 12:
                next_t = datetime.datetime(t.year + 1, 1, 1, 0, 0, 0)
            else:
                next_t = datetime.datetime(t.year, t.month + 1, 1, 0, 0, 0)
        remaining = next_t - t
        return remaining

    def calculate_bollinger_bands(prices: Series, window=20, num_std=2):

        """Calculate the Bollinger Bands."""
        rolling_mean = prices.rolling(window=window).mean()
        rolling_std = prices.rolling(window=window).std()
        upper_band = rolling_mean + (rolling_std * num_std)
        lower_band = rolling_mean - (rolling_std * num_std)
        return upper_band, lower_band

    def replace_position_with_new_limit_order(self, symbol_pair, order_id, buy_amount, new_price):

        try:
            # First, cancel the original order
            cancel_payload = {
                "symbol": symbol_pair,
                "orderId": order_id
            }
            cancel_response = self.send_signed_request("DELETE", "/api/v3/order", cancel_payload)
            std_log(f"[{symbol_pair}] Original order canceled. Info: {cancel_response}")

            new_order = self.Buy(symbol, quantity, latest_lower_bband_price)
            symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER
            return new_order

        except Exception as e:
            std_log(f"[{symbol_pair}] Error modifying order. Error Info: {e}")
            return None

    def validate_order_status(self, symbol_pair_order_status):

        # Check if all values in the dictionary are 'POSITION'
        if all(order_status == OrderStatus.POSITION for order_status in symbol_pair_order_status.values()):
            return False  # All orders are in POSITION, so return False

        return True  # Continue running-

    def set_take_profit(self, symbol_pair, buy_amount, take_profit_price, symbol_pair_target_order_count):

        payload = {
            "symbol": symbol,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "quantity": self.qtyRound(symbol, buy_amount),
            "price": self.priceRound(symbol, take_profit_price)
        }

        try:
            response = self.send_signed_request("POST", "/api/v3/order", payload)
            symbol_pair_target_order_count[symbol_pair] += 1
            return response
        except Exception as e:
            std_log(f"[{symbol_pair}] Error placing take_profit for order. Error Info: {e}")
            return None

    def update_take_profit(self, symbol_pair, order_id, new_buy_amount, new_take_profit_price,
                           symbol_pair_target_order_count):

        # Step 1: Cancel the existing order
        self.OrderCancel(symbol_pair, order_id)

        # Step 2: Place a new take profit order with the new target price
        new_payload = {
            "symbol": symbol_pair,
            "side": "SELL",
            "type": "LIMIT",
            "timeInForce": "GTC",
            "quantity": self.qtyRound(symbol_pair, new_buy_amount),
            "price": self.priceRound(symbol_pair, new_take_profit_price)
        }

        try:
            new_response = self.send_signed_request("POST", "/api/v3/order", new_payload)
            if new_response and 'orderId' in new_response:
                symbol_pair_target_order_count[symbol_pair] += 1
                std_log(f"[{symbol_pair}] New take profit order placed. Order ID: {new_response['orderId']}")
                return new_response
            else:
                std_log(f"[{symbol_pair}] Failed to place new take profit order. Response: {new_response}")
                return None
        except Exception as e:
            std_log(f"[{symbol_pair}] Error placing new take profit order. Error Info: {e}")
            return None

    def check_order_status(self, symbol_pair, order_id):

        payload = {
            "symbol": symbol_pair,
            "orderId": order_id
        }

        try:
            order_status = self.send_signed_request("GET", "/api/v3/order", payload)
            std_log(f"[{symbol_pair}] Order ID [{order_id}] with status: [{order_status['status']}]")
            return order_status['status']

        except Exception as e:
            std_log(f"[{symbol_pair}] Error checking order status. Error Info: {e}")
            return None

    def update_order_id_for_symbol_pair(self, order_info, symbol_pair):

        global symbol_pair_order_count
        if order_info is not None:
            symbol_pair_order_count[symbol_pair] = order_info['orderId']

        return symbol_pair_order_count[symbol_pair]

    def buy_and_set_take_profit(self, symbol_pair, quantity, buy_price, upper_band):

        """Buy and immediately set a take profit SELL order at the Upper Bollinger Band."""

        # Place BUY order
        buy_order = self.Buy(symbol_pair, quantity, buy_price)

        if buy_order and "orderId" in buy_order:
            # Order placed successfully, now place take profit order
            take_profit = self.set_take_profit(symbol_pair, quantity, upper_band)

            return buy_order, take_profit

        return buy_order, None

    def sell_and_set_take_profit(self, symbol_pair, quantity, sell_price, lower_band):

        """Sell and immediately set a take profit BUY order at the Lower Bollinger Band."""

        # Place SELL order
        buy_order = self.Buy(symbol_pair, quantity, sell_price)

        if buy_order and "orderId" in buy_order:
            # Order placed successfully, now place take profit order
            take_profit = self.set_take_profit(symbol_pair, quantity, lower_band)

            return buy_order, take_profit

        return buy_order, None

    def check_buy_signal(self, candles_chart_df: pd.DataFrame, symbol_pair):

        closes = candles_chart_df['Close'].values
        dates = pd.to_datetime(candles_chart_df.index)
        close_prices = pd.Series(closes, index=dates)

        # Calculate Bollinger Bands
        rolling_mean = close_prices.rolling(window=20).mean()
        rolling_std = close_prices.rolling(window=20).std()
        upper_band = rolling_mean + (2 * rolling_std)
        lower_band = rolling_mean - (2 * rolling_std)

        latest_close = closes[-1]
        latest_upper_band = upper_band.iloc[-1]
        latest_lower_band = lower_band.iloc[-1]

        # Check if the latest close price is below or equal to the Lower Bollinger Band
        if latest_close <= latest_lower_band:
            std_log(f"[{symbol_pair}] Latest Close Price {latest_close} is below or equal to "
                    f"the Lower Latest Bollinger Band {latest_lower_band}")
            return True, latest_lower_band, latest_upper_band
        else:
            std_log(f"[{symbol_pair}]  Bollinger Bands filter condition not met "
                    f"(Latest Lower Bollinger Band: {latest_lower_band} / Latest Close: {latest_close})")
            return False, latest_lower_band, latest_upper_band

    def check_sell_signal(self, candles_chart_df: pd.DataFrame, symbol_pair):

        closes = candles_chart_df['Close'].values
        dates = pd.to_datetime(candles_chart_df.index)
        close_prices = pd.Series(closes, index=dates)

        # Calculate Bollinger Bands
        rolling_mean = close_prices.rolling(window=20).mean()
        rolling_std = close_prices.rolling(window=20).std()
        upper_band = rolling_mean + (2 * rolling_std)
        lower_band = rolling_mean - (2 * rolling_std)

        latest_close = closes[-1]
        latest_upper_band = upper_band.iloc[-1]
        latest_lower_band = lower_band.iloc[-1]

        # Check if the latest close price is above or equal to the Upper Bollinger Band
        if latest_close >= latest_upper_band:
            std_log(f"[{symbol_pair}] Latest Close Price {latest_close} is above or equal to "
                    f"the Upper Latest Bollinger Band {latest_upper_band}")
            return True, latest_lower_band, latest_upper_band
        else:
            std_log(f"[{symbol_pair}]  Bollinger Bands filter condition not met "
                    f"(Latest Upper Bollinger Band: {latest_upper_band} / Latest Close: {latest_close})")
            return False, latest_lower_band, latest_upper_band

if __name__ == "__main__":

    # Secrets & Parameters 👇🔐
    CONFIG_PATH = os.getenv('CONFIG_PATH')
    EXCEL_NAME = "/Bot_config.xlsx"
    APIKEY = os.getenv('TEST_API_KEY')
    SECRETKEY = os.getenv('TEST_API_SECRET')

    # 0. Get Configuration ====================================#
    strategies = []
    buy_timeframe = "1m"
    tdelta_conv = {"1m": datetime.timedelta(minutes=1), "3m": datetime.timedelta(minutes=3),
                   "5m": datetime.timedelta(minutes=5), "15m": datetime.timedelta(minutes=15),
                   "30m": datetime.timedelta(minutes=30), "1h": datetime.timedelta(hours=1),
                   "2h": datetime.timedelta(hours=2), "4h": datetime.timedelta(hours=4),
                   "6h": datetime.timedelta(hours=6), "8h": datetime.timedelta(hours=8),
                   "12h": datetime.timedelta(hours=12), "1d": datetime.timedelta(days=1),
                   "3d": datetime.timedelta(days=3), "1w": datetime.timedelta(days=7)}
    set_filled = [False, ] * 6


    buy_timeframe = {}
    buy_order_type = {}
    sell_timeframe = {}
    sell_order_type = {}
    order_size = {}
    h_period = {}
    l_period = {}
    buy_limit = 0

    # 0.1. Read excel
    config = openpyxl.load_workbook(CONFIG_PATH + EXCEL_NAME)
    sheet = config.worksheets[0]

    for row_n in range(2, 100):
        if sheet.cell(row=row_n, column=1).value == "END OF CONFIGURATION":
            break
        if sheet.cell(row=row_n, column=3).value == "BUY":
            symbol = sheet.cell(row=row_n, column=1).value
            if not symbol in symbols:
                symbols.append(symbol)
            #currency = sheet.cell(row=row_n, column=2).value
            buy_timeframe[symbol] = sheet.cell(row=row_n, column=4).value
            buy_order_type[symbol] = sheet.cell(row=row_n, column=5).value
            order_size[symbol] = float(sheet.cell(row=row_n, column=6).value)
            h_period[symbol] = int(sheet.cell(row=row_n, column=7).value)
        elif sheet.cell(row=row_n, column=3).value == "SELL":
            sell_order_type[symbol] = sheet.cell(row=row_n, column=5).value
            sell_timeframe[symbol] = sheet.cell(row=row_n, column=4).value
            l_period[symbol] = int(sheet.cell(row=row_n, column=8).value)
        elif sheet.cell(row=row_n, column=1).value == "Open positions limit":
            buy_limit = int(sheet.cell(row=row_n, column=2).value)
        elif sheet.cell(row=row_n, column=1).value == "Demo trading":
            demo = sheet.cell(row=row_n, column=2).value
        elif sheet.cell(row=row_n, column=1).value == "EMA_filter":
            ema_period = int(sheet.cell(row=row_n, column=2).value)
        elif sheet.cell(row=row_n, column=1).value == "EMA_timeframe":
            ema_timeframe = sheet.cell(row=row_n, column=2).value
    #std_log("[Booting]%s"%str(
    std_log("[Booting] Buy timeframes %s" % str(buy_timeframe))
    std_log("[Booting] Buy order type %s" % str(buy_order_type))
    std_log("[Booting] Sell timeframes %s" % str(sell_timeframe))
    std_log("[Booting] Sell order type %s" % str(sell_order_type))
    std_log("[Booting] Order size %s" % str(order_size))
    std_log("[Booting] Highest period %s" % str(h_period))
    std_log("[Booting] Lowest period %s" % str(l_period))
    std_log("[Booting] Open positions limit %s" % str(buy_limit))
    std_log("[Booting] Demo account: %s" % str(demo))

    buy_timedelta = {}
    sell_timedelta = {}
    for symbol in symbols:
        buy_timedelta[symbol] = tdelta_conv[buy_timeframe[symbol]]  # to get candle closing period
        sell_timedelta[symbol] = tdelta_conv[sell_timeframe[symbol]]
    ema_timedelta = tdelta_conv[ema_timeframe]

    # 1. API init
    binance = Binance(test=demo, apikey=APIKEY, secretkey=SECRETKEY)

    # 2. Balance recovery
    balance = []
    if not os.path.isfile(CONFIG_PATH + "balance.txt"):
        fout = open(CONFIG_PATH + "balance.txt", "w")
        fout.writelines("[]")
        fout.close()
    else:
        try:
            fin = open(CONFIG_PATH + "balance.txt", "r")
            balance = eval(fin.readline())
            fin.close()
            for b in balance:
                if type(b) != dict:
                    raise TypeError
            if len(balance) != 0:
                std_log("[Booting] Balance recovered: %s" % str(balance))
        except:
            std_log("[Booting] balance.txt broken, reset balance.")
            fout = open(CONFIG_PATH + "balance.txt", "w")
            fout.writelines("[]")
            fout.close()
            # {"symbol":symbol,"quantity":0,"entry_t":0,"orderId":""}

    std_log("[Booting] Complete")

    # 3. Trading routine
    old_remain_buy = {}
    old_remain_sell = {}
    for symbol in symbols:
        old_remain_buy[symbol] = datetime.timedelta(days=365)  # to get candle closing period
        old_remain_sell[symbol] = datetime.timedelta(days=365)  # to get candle closing period

    symbol_pair_trade_counter = {symbol: 0 for symbol in symbols}
    symbol_pair_order_count = {symbol: 0 for symbol in symbols}
    symbol_pair_target_order_count = {symbol: 0 for symbol in symbols}
    symbol_pair_order_status = {symbol: OrderStatus.OPEN_ORDER for symbol in symbols}
    symbol_pair_target_order_status = {symbol: OrderStatus.OPEN_ORDER for symbol in symbols}

    while True:
        showed_remain = datetime.timedelta(days=365)  # To show remain time to candle closing

        for symbol in symbols:
            # 3.0. Check if balance has same position
            exist_in_balance = False
            balance_idx = -1
            for bi in range(len(balance)):
                b = balance[bi]
                if b["symbol"] == symbol:
                    exist_in_balance = True
                    balance_idx = bi

            # 3.1. Buy Routine
            if len(balance) < buy_limit and (not exist_in_balance):  # Check if balance is full

                remain = binance.boundaryRemaining(buy_timeframe[symbol])  # Remain time to buy candle closing
                showed_remain = min(remain, showed_remain)

                if old_remain_buy[symbol] < remain:  # Get into new candle

                    chart, chart_df = binance.getChart(symbol,
                                                       buy_timeframe[symbol],
                                                       start_t=datetime.datetime.now(datetime.UTC) - buy_timedelta[symbol]
                                                               * h_period[symbol] * 2)

                    if chart["t"][-1] > datetime.datetime.now(datetime.UTC) - buy_timedelta[symbol] / 2:
                        chart["t"] = chart["t"][:-1]  # Remove excessive candle
                        chart["o"] = chart["o"][:-1]
                        chart["h"] = chart["h"][:-1]
                        chart["l"] = chart["l"][:-1]
                        chart["c"] = chart["c"][:-1]
                        chart["v"] = chart["v"][:-1]

                    high_hit = chart["c"][-1] >= max(chart["c"][-h_period[symbol] - 1:-1])

                    # Ensure DataFrame is not empty and is sorted by index (Open Time)
                    if chart_df.empty or chart_df.index[-1] > datetime.datetime.now(datetime.UTC) - sell_timedelta[symbol] / 2:
                        # Remove the last row if the candle is not fully formed
                        chart_df = chart_df.iloc[:-1]

                    bbands_ok, latest_lower_bband_price, latest_upper_bband_price = binance.check_buy_signal(
                        chart_df, symbol)

                    close_p = chart["c"][-1]
                    quantity = order_size[symbol] / close_p


                    if symbol_pair_order_count[symbol] > 0:

                            position_status = binance.check_order_status(symbol, symbol_pair_target_order_count[symbol])
                            order_id = symbol_pair_order_count[symbol]

                            if position_status == 'FILLED':

                                take_profit_response = binance.set_take_profit(symbol,
                                                                               quantity,
                                                                               latest_upper_bband_price,
                                                                               symbol_pair_target_order_count)

                                # Update balance if order filled
                                asset = binance.getBase(symbol)
                                quantity = min(quantity,
                                               binance.getBalanceQuantity(asset))  # To check real amount in balance
                                price = binance.GetContractPrice(symbol, order_id)
                                std_log("%s %g bought" % (symbol, price))
                                balance.append(
                                    {"symbol": symbol, "quantity": quantity, "entry_t": datetime.datetime.now(),
                                     "orderId": order_id, "buyprice": price})
                                update_balance(balance)
                                if len(balance) == buy_limit:
                                    std_log("Balance is full")

                            else:

                                new_order_response = binance.replace_position_with_new_limit_order(symbol,
                                                                              order_id,
                                                                              quantity,
                                                                              latest_lower_bband_price)


                            if (position_status == 'FILLED'
                                    and symbol_pair_target_order_count[symbol] > 0):

                                take_profit_status = binance.check_order_status(symbol, symbol_pair_target_order_count[symbol])

                                if take_profit_status != 'FILLED':
                                    new_take_profit_order = binance.update_take_profit(symbol,
                                                                                       symbol_pair_target_order_count[symbol],
                                                                                       quantity,
                                                                                       latest_upper_bband_price,
                                                                                       symbol_pair_target_order_count)
                                else:
                                    symbol_pair_order_count[symbol] = 0
                                    symbol_pair_order_status[symbol] = OrderStatus.OPEN_ORDER
                                    symbol_pair_target_order_count[symbol] = 0
                                    symbol_pair_target_order_status[symbol] = OrderStatus.OPEN_ORDER


                    if high_hit:

                        if bbands_ok:  # Buy condition met

                            if buy_order_type[symbol] == "LMT":
                                result = binance.Buy(symbol, quantity, latest_lower_bband_price)
                            else:
                                # Else submitting MARKET order at current price
                                result = binance.Buy(symbol, quantity, 0)

                            binance.update_order_id_for_symbol_pair(result, symbol)

                        else:
                            print("Bollinger Band Condition Not Met. No Order/Positions Set. "
                                  "Check logs for more information.")
                    else:
                        std_log("[%s]  Highest close does not met" % (symbol,))

                old_remain_buy[symbol] = remain

            # 3.2. Sell Routine
            elif exist_in_balance:  # Check if balance has any position
                remain = binance.boundaryRemaining(sell_timeframe[symbol])
                showed_remain = min(remain, showed_remain)  # Remain time to sell candle closing
                if old_remain_sell[symbol] < remain:  # Get into new candle

                    chart, chart_df = binance.getChart(symbol,
                                                       buy_timeframe[symbol],
                                                       start_t=datetime.datetime.now(datetime.UTC) - buy_timedelta[symbol]
                                                               * h_period[symbol] * 2)

                    if chart["t"][-1] > datetime.datetime.now(datetime.UTC) - sell_timedelta[symbol] / 2:
                        chart["t"] = chart["t"][:-1]  # Remove excessive candle
                        chart["o"] = chart["o"][:-1]
                        chart["h"] = chart["h"][:-1]
                        chart["l"] = chart["l"][:-1]
                        chart["c"] = chart["c"][:-1]
                        chart["v"] = chart["v"][:-1]

                    low_hit = chart["c"][-1] <= min(chart["c"][-l_period[symbol] - 1:-1])

                    std_log("[%s] Bar close:%g / Lowest:%g" % (symbol,
                                                               chart["c"][-1],
                                                               min(chart["c"][-l_period[symbol] - 1:-1])))

                    # Ensure DataFrame is not empty and is sorted by index (Open Time)
                    if chart_df.empty or chart_df.index[-1] > datetime.datetime.now(datetime.UTC) - sell_timedelta[symbol] / 2:
                        # Remove the last row if the candle is not fully formed
                        chart_df = chart_df.iloc[:-1]

                    bbands_ok, latest_lower_bband_price, latest_upper_bband_price = binance.check_sell_signal(
                        chart_df, symbol)

                    if low_hit:  # Sell condition met

                        if bbands_ok:

                            close_p = chart["c"][-1]
                            asset = binance.getBase(symbol)
                            balance[balance_idx]["quantity"] = min(balance[balance_idx]["quantity"],
                                                                   binance.getBalanceQuantity(asset))  # Feb15 added
                            quantity = binance.qtyRoundDown(symbol, balance[balance_idx]["quantity"])

                            std_log("%s sell (qty: %g)" % (symbol, quantity))

                            if sell_order_type[symbol] == "LMT":

                                # Placing a LIMIT SELL Order at the Latest Upper Bollinger Band Price
                                # & TAKE_PROFIT on the Latest Lower Bollinger Band Price
                                (result, take_profit_response) = binance.sell_and_set_take_profit(symbol, quantity,
                                                                                                  latest_upper_bband_price,
                                                                                                  latest_lower_bband_price)
                            else:

                                # Placing a MARKET BUY Order at the current price
                                result = binance.Sell(symbol, quantity, 0)

                            orderId = result["orderId"]
                            order_filled = binance.OrderWait(symbol, orderId)

                            if order_filled:
                                # Update balance if order filled
                                price = binance.GetContractPrice(symbol, orderId)
                                std_log("%s %g sold" % (symbol, price))
                                del balance[balance_idx]
                                update_balance(balance)

                        else:
                            print("Bollinger Band Condition Not Met. No Order/Positions Set. "
                                  "Check logs for more information.")
                    else:
                        std_log("[%s]  Lowest close does not met" % (symbol,))

                old_remain_sell[symbol] = remain

        # 3.3. Log
        r_hour = int(showed_remain.seconds / 3600)
        r_minute = int((showed_remain.seconds - 3600 * r_hour) / 60)
        r_second = showed_remain.seconds - 3600 * r_hour - 60 * r_minute
        text = ""

        if showed_remain.days == 0:
            text = text + " Time to next candle .. %02d:%02d:%02d (balance: %d/%d)           " % (r_hour,
                                                                                                  r_minute, r_second,
                                                                                                  len(balance),
                                                                                                  buy_limit)
        else:
            text = text + " Time to next candle .. %d days %02d:%02d:%02d (balance: %d/%d)   " % (showed_remain.days,
                                                                                                  r_hour, r_minute,
                                                                                                  r_second,
                                                                                                  len(balance),
                                                                                                  buy_limit)

        print("\r" + cur_time() + text, end="\r")
        time.sleep(1)
